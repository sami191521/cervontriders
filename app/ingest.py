"""Five9 export -> normalized rider dicts (handoff §3).

Mirrors the HTML reference (CANON aliases + applyMapping):
- parse .xlsx (openpyxl) or .csv/.tsv
- auto-guess header -> field mapping, case/whitespace-insensitive
- normalize: strip %/comma, blank->0, CR- prefix, pp/validpct decimals ->0-100,
  keep target/conv as decimals, round integer fields.
"""
import io
import re
import csv as csvmod
from typing import Optional

# canonical fields + the header aliases they match (HTML CANON / handoff §3)
# (key, label, required, [aliases])
CANON = [
    ("n",        "Agent Name",                     True,  ["agent name", "name", "agent"]),
    ("tlraw",    "Team Leader",                     False, ["team leader", "supervisor", "team"]),
    ("tn",       "Tenure",                          False, ["tenure range", "tenure", "actual tenure"]),
    ("status",   "Status",                          False, ["status"]),
    ("d",        "Present Days",                    False, ["present days", "days present"]),
    ("th",       "Total Productive Hours",          False, ["total productive hours"]),
    ("ah",       "Avg Productive Hours",            False, ["avg productive hours", "average productive hours"]),
    ("pp",       "Productive %",                    False, ["productive", "productivity"]),
    ("tw",       "Total Wait Time",                 False, ["total wait time"]),
    ("aw",       "Avg Wait Time",                   False, ["avg wait time", "average wait time"]),
    ("ld",       "Sales (Five9 Lead Submitted)",    True,  ["five9 lead submitted", "lead submitted", "leads", "sales", "submitted"]),
    ("r",        "Sales / day (Submitted Ratio)",   False, ["five9 submitted ratio", "submitted ratio", "sales per day", "ratio"]),
    ("target",   "Target Achieved %",               False, ["target achieved", "target"]),
    ("conv",     "Conversion %",                    False, ["connect conversion rate", "convertion rate", "conversion rate", "conversion"]),
    ("valid",    "Valid Lead Count",                False, ["valid lead count", "valid leads"]),
    ("validpct", "Valid %",                         False, ["valid %", "valid percent", "valid pct"]),
    ("mva",      "MVA Leads",                       False, ["mva lead submitted", "mva"]),
    ("wc",       "Workers Comp Leads",              False, ["worker compensation", "workers compensation", "worker comp"]),
    ("pi",       "Personal Injury Leads",           False, ["personal injury"]),
    ("oth",      "Other Leads",                     False, ["others", "other"]),
]

# fields treated as integers (rounded)
_INT_FIELDS = {"ld", "valid", "mva", "wc", "pi", "oth"}
# fields that are decimals in-file and shown ×100 by the UI -> keep as-is
_DECIMAL_FIELDS = {"target", "conv"}
# optional fields: only set when the column was actually mapped, else None
_OPTIONAL_FIELDS = {"target", "conv", "valid", "validpct", "mva", "wc", "pi", "oth"}


def norm(s) -> str:
    """Lowercase, collapse non-alphanumerics to single spaces, trim (HTML norm)."""
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def slugify(name: str) -> str:
    """Stable id, identical to the front-end (no leading/trailing trim)."""
    return re.sub(r"[^a-z0-9]+", "-", str(name).lower())


def to_num(v) -> float:
    """Parse a numeric cell: blank/(Blank)->0, strip % and commas, NaN->0."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if s == "" or re.search(r"blank", s, re.I):
        return 0.0
    s = re.sub(r"[%,]", "", s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def guess_header(headers: list[str], aliases: list[str]) -> str:
    """Best header match for a field: exact normalized match, then substring."""
    pairs = [(h, norm(h)) for h in headers]
    for al in aliases:
        a = norm(al)
        for h, hn in pairs:
            if hn == a:
                return h
        for h, hn in pairs:
            if a and a in hn:
                return h
    return ""


def auto_map(headers: list[str]) -> dict:
    """Auto-guess {field_key -> header}. Admin can override before applying."""
    return {key: guess_header(headers, aliases) for (key, _lbl, _req, aliases) in CANON}


# ---------- parsing ----------------------------------------------------------
def parse_delimited(text: str) -> tuple[list[str], list[dict]]:
    """CSV/TSV -> (headers, rows). Sniffs tab vs comma from the header line."""
    text = text.replace("\r", "").strip()
    if not text:
        return [], []
    first = text.split("\n", 1)[0]
    delim = "\t" if "\t" in first else ","
    reader = csvmod.reader(io.StringIO(text), delimiter=delim)
    rows_raw = [r for r in reader]
    if not rows_raw:
        return [], []
    headers = [h.strip() for h in rows_raw[0]]
    rows = []
    for cells in rows_raw[1:]:
        if not any(str(c).strip() for c in cells):
            continue
        rows.append({h: (str(cells[i]).strip() if i < len(cells) else "")
                     for i, h in enumerate(headers)})
    return headers, rows


def parse_xlsx(data: bytes) -> tuple[list[str], list[dict]]:
    """First sheet of an .xlsx -> (headers, rows)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return [], []
    headers = [("" if h is None else str(h).strip()) for h in header_row]
    rows = []
    for raw in it:
        if raw is None or not any(c not in (None, "") for c in raw):
            continue
        rows.append({headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)})
    wb.close()
    return headers, rows


def parse_upload(filename: str, data: bytes) -> tuple[list[str], list[dict]]:
    """Dispatch on extension. .xlsx/.xls -> openpyxl, else delimited text."""
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext in ("xlsx", "xls"):
        return parse_xlsx(data)
    return parse_delimited(data.decode("utf-8", errors="replace"))


# ---------- normalization ----------------------------------------------------
def _pct(v: float) -> float:
    """A 0-100 percentage: values <=2 are fractions (0.98) -> ×100."""
    return v * 100 if (v and v <= 2) else v


def build_riders(rows: list[dict], field_map: dict) -> list[dict]:
    """Apply the mapping + normalization rules; return plain rider dicts.

    Raises ValueError if the required name/sales fields aren't mapped.
    """
    if not field_map.get("n") or not field_map.get("ld"):
        raise ValueError("Map at least Agent Name and Sales.")

    def col(row, key):
        h = field_map.get(key)
        return row.get(h) if h else None

    out = []
    for row in rows:
        name = str(col(row, "n") or "").strip()
        if not name:
            continue

        pp = _pct(to_num(col(row, "pp")))
        tlraw = str(col(row, "tlraw") or "").strip()
        tl = re.sub(r"^cr\s*-?\s*", "", tlraw, flags=re.I).strip() or "Unassigned"

        rider = {
            "n": name,
            "id": slugify(name),
            "tl": tl,
            "tlraw": tlraw,
            "tn": str(col(row, "tn") or "").strip(),
            "status": str(col(row, "status") or "").strip(),
            "d": to_num(col(row, "d")),
            "th": to_num(col(row, "th")),
            "ah": to_num(col(row, "ah")),
            "pp": pp,
            "tw": to_num(col(row, "tw")),
            "aw": to_num(col(row, "aw")),
            "ld": round(to_num(col(row, "ld"))),
            "r": to_num(col(row, "r")),
        }

        # optional fields: None unless the column was mapped
        for key in _OPTIONAL_FIELDS:
            if not field_map.get(key):
                rider[key] = None
                continue
            v = to_num(col(row, key))
            if key == "validpct":
                rider[key] = _pct(v)
            elif key in _DECIMAL_FIELDS:
                rider[key] = v
            elif key in _INT_FIELDS:
                rider[key] = round(v)
            else:
                rider[key] = v
        out.append(rider)

    if not out:
        raise ValueError("No rows parsed — check the mapping.")
    return out
